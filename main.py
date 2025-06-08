from telegram import Update
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    filters,
    ContextTypes,
)
import logging
from pydub import AudioSegment
import whisper
import re
import os
from thefuzz import fuzz
from pytz import timezone
import random
from openpyxl import load_workbook
from typing import List, Tuple
from excel_manager import ExcelManager
import openpyxl
from dotenv import load_dotenv
import os

load_dotenv()  # загружает переменные из .env

TOKEN = os.getenv("BOT_TOKEN")


# --- Класс ExcelManager (твоя версия) ---

class ExcelManager:
    def __init__(self, file_path: str = "tenderisthenight.xlsx"):
        self.file_path = file_path
        self.wb = load_workbook(filename=self.file_path)
        self.sheet = self.wb.active

    def get_all_pairs(self) -> List[Tuple[str, str, str]]:
        # Читаем три колонки: English, IPA, Russian
        rows = list(self.sheet.iter_rows(min_row=2, max_col=3, values_only=True))
        filtered = [
            (eng, ipa or "", rus) for eng, ipa, rus in rows
            if eng and eng.strip() != "" and rus and rus.strip() != ""
        ]
        return filtered

    def get_random_pair(self) -> Tuple[str, str, str]:
        pairs = self.get_all_pairs()
        if not pairs:
            raise ValueError("Файл пуст или все строки пустые!")
        return random.choice(pairs)

    def get_random_words(self, n: int = 10) -> List[str]:
        all_english_words = [row[0] for row in self.get_all_pairs()]
        return random.sample(all_english_words, min(len(all_english_words), n))

    def add_phrase(self, english: str, ipa: str, russian: str):
        self.sheet.append([english, ipa, russian])
        self.wb.save(self.file_path)


# --- Конфигурация и загрузка модели ---
desktop_cache = r"C:\Users\PC\Desktop\.cache"
os.makedirs(desktop_cache, exist_ok=True)
os.environ["XDG_CACHE_HOME"] = desktop_cache

model = whisper.load_model("base")

logging.basicConfig(format="%(asctime)s - %(name)s - %(levelname)s - %(message)s", level=logging.INFO)
logger = logging.getLogger(__name__)

# --- Создаем менеджер ---
excel_manager = ExcelManager("tenderisthenight.xlsx")

# --- Вспомогательные функции ---
def normalize_text(text):
    text = text.lower()
    text = re.sub(r'[^a-z\s]', '', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text

def word_by_word_analysis(correct_text, predicted_text, threshold=0.6):
    correct_words = normalize_text(correct_text).split()
    predicted_words = normalize_text(predicted_text).split()

    analysis = []
    used_indices = set()

    for i, cw in enumerate(correct_words):
        match = None
        best_score = 0
        best_word = None

        for j, pw in enumerate(predicted_words):
            if j in used_indices:
                continue
            score = fuzz.ratio(cw, pw) / 100
            if score > best_score:
                best_score = score
                best_word = pw
                match = (pw, score, j)

        if best_score >= threshold:
            analysis.append(f"✅ {cw}")
            used_indices.add(match[2])
        elif best_word:
            analysis.append(f"❌ {cw} ➜ вы сказали: {best_word} ({int(best_score * 100)}%)")
            used_indices.add(match[2])
        else:
            analysis.append(f"❗ Пропущено слово: {cw}")

    return analysis

# --- Обработчики команд ---
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    eng, ipa, rus = excel_manager.get_random_pair()
    context.user_data["current_phrase"] = eng

    await update.message.reply_text(
        f"Привет, {user.first_name}! 👋\n"
        f"Давайте потренируем ваше произношение:\n\n"
        f"🇬🇧 Текст: <b>{eng}</b>\n"
        f"🗣 Транскрипция: <code>{ipa}</code>\n"
        f"🇷🇺 Перевод: <i>{rus}</i>\n\n"
        "Отправьте голосовое сообщение с вашим произношением!",
        parse_mode="HTML"
    )
    

async def next_phrase(update: Update, context: ContextTypes.DEFAULT_TYPE):
    eng, ipa, rus = excel_manager.get_random_pair()
    ipa = ""

    context.user_data["current_phrase"] = eng

    await update.message.reply_text(
        f"🇬🇧 Следующая фраза: <b>{eng}</b>\n"
        f"🗣 Транскрипция: <code>{ipa}</code>\n"
        f"🇷🇺 Перевод: <i>{rus}</i>",
        parse_mode="HTML"
    )

async def words_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    random_words = excel_manager.get_random_words()
    word_list = '\n'.join(random_words)
    await update.message.reply_text(
        f"Список случайных слов для тренировки:\n\n<b>{word_list}</b>",
        parse_mode="HTML"
    )

async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "📋 Вот, что я умею:\n\n"
        "/start — начать тренировку с новой фразой\n"
        "/next — следующая фраза\n"
        "/words — случайные слова для тренировки\n"
        "🎤 Отправь голосовое сообщение, чтобы я проверил произношение!"
    )

# --- Обработка голосовых сообщений ---
async def handle_voice(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if "current_phrase" not in context.user_data:
        await update.message.reply_text("Сначала используйте /start")
        return

    correct_phrase = context.user_data["current_phrase"]

    voice_file = await update.message.voice.get_file()
    input_ogg = "input.ogg"
    input_wav = "input.wav"

    try:
        await voice_file.download_to_drive(input_ogg)
        sound = AudioSegment.from_file(input_ogg)
        sound.export(input_wav, format="wav")

        result = model.transcribe(input_wav, language="en")
        transcription = result["text"].strip()
        logger.info(f"Transcription: {transcription}")

        word_results = word_by_word_analysis(correct_phrase, transcription)
        feedback = "\n".join(word_results)

        await update.message.reply_text(
            f"<b>Разбор по словам:</b>\n\n{feedback}",
            parse_mode="HTML"
        )

    except Exception as e:
        logger.error(f"Ошибка при обработке голосового сообщения: {e}")
        await update.message.reply_text("Произошла ошибка при распознавании речи. Попробуйте ещё раз.")
    finally:
        if os.path.exists(input_ogg):
            os.remove(input_ogg)
        if os.path.exists(input_wav):
            os.remove(input_wav)

# --- Запуск бота ---
def main():
    application = Application.builder().token("YOUR_TOKEN").build()
    application.job_queue.scheduler.configure(timezone=timezone("Europe/Moscow"))

    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("next", next_phrase))
    application.add_handler(CommandHandler("words", words_command))
    application.add_handler(CommandHandler("help", help_command))
    application.add_handler(MessageHandler(filters.VOICE, handle_voice))

    application.run_polling()

if __name__ == "__main__":
    main()
