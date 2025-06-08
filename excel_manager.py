from openpyxl import load_workbook
from typing import List, Tuple
import random
import logging
from telegram import Update
from telegram.ext import ContextTypes

# Настройка логгера
logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO)

class ExcelManager:
    def __init__(self, file_path: str = "tenderisthenight.xlsx"):
        self.file_path = file_path
        self.wb = load_workbook(filename=self.file_path)
        self.sheet = self.wb.active

    def get_all_pairs(self) -> List[Tuple[str, str, str]]:
        # Читаем три колонки: English, IPA, Russian
        rows = list(self.sheet.iter_rows(min_row=2, max_col=3, values_only=True))
        filtered = [
            (eng, ipa or "", rus) for eng, ipa, rus in rows
            if eng and eng.strip() != "" and rus and rus.strip() != ""
        ]
        return filtered

    def get_random_pair(self) -> Tuple[str, str, str]:
        pairs = self.get_all_pairs()
        if not pairs:
            raise ValueError("Файл пуст или все строки пустые!")
        return random.choice(pairs)

    def get_random_words(self, n: int = 10) -> List[str]:
        all_english_words = [row[0] for row in self.get_all_pairs()]
        return random.sample(all_english_words, min(len(all_english_words), n))

    def add_phrase(self, english: str, ipa: str, russian: str):
        self.sheet.append([english, ipa, russian])
        self.wb.save(self.file_path)


# Создаём экземпляр менеджера
excel_manager = ExcelManager("tenderisthenight.xlsx")

# Команда /start
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    logger.info(f"User {user.id} started the bot")

    eng, ipa, rus = excel_manager.get_random_pair()
    context.user_data["current_phrase"] = eng

    await update.message.reply_text(
        f"Привет, {user.first_name}! 👋\n"
        f"Давайте потренируем ваше произношение:\n\n"
        f"🇬🇧 Текст: <b>{eng}</b>\n"
        f"🗣 Транскрипция: <code>{ipa}</code>\n"
        f"🇷🇺 Перевод: <i>{rus}</i>\n\n"
        "Отправьте голосовое сообщение с вашим произношением!",
        parse_mode="HTML"
    )

# Команда /next
async def next_phrase(update: Update, context: ContextTypes.DEFAULT_TYPE):
    eng, ipa, rus = excel_manager.get_random_pair()
    context.user_data["current_phrase"] = eng

    await update.message.reply_text(
        f"🇬🇧 Следующая фраза: <b>{eng}</b>\n"
        f"🗣 Транскрипция: <code>{ipa}</code>\n"
        f"🇷🇺 Перевод: <i>{rus}</i>",
        parse_mode="HTML"
    )

# Команда /words
async def words_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    random_words = excel_manager.get_random_words()
    word_list = '\n'.join(f"• {word}" for word in random_words)
    await update.message.reply_text(
        f"Список случайных слов для тренировки:\n\n<b>{word_list}</b>",
        parse_mode="HTML"
    )
